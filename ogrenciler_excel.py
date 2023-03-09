import tkinter as tk
from tkinter import ttk
import openpyxl as xl
import re

"""
'ogr.xlsx' adli dosya ile aynı konumda
olmasına dikkat ediniz
"""

window = tk.Tk()
arka_plan = "#4A66F0"
window.configure(bg=arka_plan)
window.geometry("400x250+500+200")
window.title("Ogrenciler")
window.resizable(height=False,width=False)
lab = tk.Label(text="Created by Efe Kocak",bg=window["bg"]).place(x=10,y=220)
ogrenciler = []

class ogrenci():
    sayim = 0
    def __init__(self,isim,sinif = str,puan_durum = ""):
        self.isim = isim.upper()
        self.sinif = sinif.upper()
        self.puan = puan_durum
        ogrenciler.append(self)
        sayi = ogrenci.sayim 
        self.sayi = sayi
        ogrenci.sayim += 1

oku = xl.load_workbook("ogr.xlsx")
aktif = oku.active

def ogr_yaz():
    global isimler
    global soy_isim
    global ogr_sinif
    global puan
    for u in range(1,aktif.max_row+1):
        isimler = aktif[f"A{u}"].value
        soy_isim = aktif[f"B{u}"].value
        puan = aktif[f"C{u}"].value
        ogr_sinif = aktif[f"D{u}"].value
        if puan == None:
            aktif.cell(row=u,column=3,value='"')
            puan = aktif[f"C{u}"].value
            oku.save("ogr.xlsx")
        sahis = ogrenci(f"{isimler} {soy_isim}",ogr_sinif,puan)

ogr_yaz()

def sec():
    girdi = ogrenci_list.curselection() 
    if girdi:
        index = girdi[0]
        secilen = ogrenci_list.get(index)
        return secilen  

siniflar = []

for u in ogrenciler:
    if u.sinif not in siniflar:
        siniflar.append(u.sinif)


def sirala( l ): 
    """ Sort the given iterable in the way that humans expect.""" 
    convert = lambda text: int(text) if text.isdigit() else text 
    alphanum_key = lambda key: [ convert(c) for c in re.split('([0-9]+)', key) ] 
    return sorted(l, key = alphanum_key)

siniflar = sirala(siniflar)

secilen_sinif = []


def ogrenci_al():
    ogrenci_list.delete(0,tk.END)
    secilen_sinif.clear()
    al = sinif_combo_box.get()
    for u in ogrenciler:
        if u.sinif == al:
            secilen_sinif.append(u)
            isimler = [u.isim for u in secilen_sinif]
            ogrenci_list.insert(tk.END,u.isim)

lab1 = tk.Label(window,text="Sınıf",fg="black",font="bold 12",bg=arka_plan).place(x=15,y=5)

sinif_combo_box = ttk.Combobox(values=siniflar,width=9,font="bold 10")
sinif_combo_box.place(x=10,y=30)


def excel_yaz_arti(degerr = "+"):
    secilen_ogr = sec()
    if secilen_ogr != None:
        secilen_ogr = sec().lower()
    for u in range(1,aktif.max_row+1):
        isimler = aktif[f"A{u}"].value.lower()
        soy_isim = aktif[f"B{u}"].value.lower()
        puan = f'{aktif[f"C{u}"].value}'
        ogr_sinif = aktif[f"D{u}"].value.upper()
        if isimler+" "+soy_isim == secilen_ogr and sinif_combo_box.get() == ogr_sinif:
            deger = "".join(puan+degerr)
            aktif.cell(column=3,row=u,value=deger)
            oku.save("ogr.xlsx")
            for i in ogrenciler:
                if i.isim.lower() == isimler+" "+soy_isim and i.sinif == sinif_combo_box.get():
                    i.puan = "".join(degerr+i.puan)
                    break
            break
           
def excel_yaz_eksi():
    excel_yaz_arti(degerr="-")
def excel_yaz_yarim():
    excel_yaz_arti(degerr="﬩")

fontayar = ("arial",12,"bold")

def bilgi():
    global arti_sayi
    global f
    global b
    global c
    global eksi_sayi
    global yarim_sayi
    arti_sayi = 0
    eksi_sayi = 0
    yarim_sayi = 0
    yellow = "#7760EC"
    secilen_ogr = sec()
    for a in ogrenciler:
        if a.isim == secilen_ogr:
            if secilen_ogr != None:
                wd = tk.Tk()
                wd.config(bg=yellow)
                wd.resizable(height=False,width=False)
                wd.geometry("200x150+900+200")
                entr_isim = tk.Entry(wd,font=fontayar,width=20,justify=tk.CENTER)
                entr_isim.insert(0,secilen_ogr)
                entr_isim.place(x=10,y=40)
                lab_sinif = tk.Label(wd,text=a.sinif,font=("arial",14,"bold"),bg=yellow).place(x=80,y=5)
                lab_arti = tk.Label(wd,text="+",font=("arial",14,"bold"),bg=yellow).place(x=30,y=70)
                lab_yarim = tk.Label(wd,text="﬩",font=("arial",14,"bold"),bg=yellow).place(x=90,y=70)
                lab_yarim = tk.Label(wd,text="-",font=("arial",14,"bold"),bg=yellow).place(x=150,y=70)
                for u in range(1,aktif.max_row+1):
                    isimler = aktif[f"A{u}"].value.upper()
                    soy_isim = aktif[f"B{u}"].value.upper()
                    ogr_sinif = aktif[f"D{u}"].value.upper()
                    puan = aktif[f"C{u}"].value.upper()
                    if isimler+" "+soy_isim == secilen_ogr and ogr_sinif == a.sinif:
                        for a in puan:
                            if a == "+":
                                arti_sayi += 1
                            if a == "-":
                                eksi_sayi+=1
                            if a == "﬩":
                                yarim_sayi+=1
                        break
                def yaz_a(belirgec = "+"):
                    global arti_sayi
                    global eksi_sayi
                    global yarim_sayi
                    if belirgec == "+":
                        excel_yaz_arti()
                        arti_sayi += 1
                        f.config(text=arti_sayi)
                    if belirgec == "-":
                        excel_yaz_eksi()
                        eksi_sayi += 1
                        b.config(text=eksi_sayi)
                    if belirgec == "f":
                        excel_yaz_yarim()
                        yarim_sayi += 1
                        c.config(text=yarim_sayi)
                def yaz_b():
                    yaz_a(belirgec="-")
                def yaz_c():
                    yaz_a(belirgec="f")
                f=tk.Button(wd,text=arti_sayi,font="bold 12",width=5,command=yaz_a)
                f.place(x=10,y=100)
                b=tk.Button(wd,text=yarim_sayi,font="bold 12",width=5,command=yaz_b)
                b.place(x=70,y=100)
                c=tk.Button(wd,text=eksi_sayi,font="bold 12",width=5,command=yaz_c)
                c.place(x=130,y=100)
                break
ogrenci_sec = tk.Button(window,text="SEC",command=ogrenci_al,width=10,font="bold 10")
ogrenci_sec.place(x = 10,y=60)

arti_dugme = tk.Button(window,text="+",font="bold 14" ,width=4,name="arti",command=excel_yaz_arti)
# print(arti_dugme.winfo_name() == "arti")
arti_dugme.place(x=130,y=120)

yarim_arti = tk.Button(window,text="﬩",font="bold 14",width=4,command=excel_yaz_yarim,name="yarim")
yarim_arti.place(x= 70,y=120)

eksi = tk.Button(window,text="-",font="bold 14",width=4,name="eksi",command=excel_yaz_eksi)
eksi.place(x=10,y=120)

def duzenler():
    secilen_ogr = sec()
    for a in ogrenciler:
        if a.isim == secilen_ogr and a.sinif == sinif_combo_box.get():
            yellow = "#F3EA2B"
            wd = tk.Tk()
            wd.config(bg=yellow)
            wd.resizable(height=False,width=False)
            wd.geometry("170x150+900+200")
            lab_sinif = tk.Label(wd,text=a.sinif,font=("arial",14,"bold"),bg=yellow).place(x=60,y=5)
            etiket = tk.Entry(wd,font="bold 12",width=18,justify=tk.CENTER)
            etiket.insert(0,a.isim)
            etiket.place(x=1,y=40)
            entr = tk.Entry(wd,font="bold 12",width=12)
            entr.insert(0,f'"{a.puan}')
            entr.place(x=20,y=70)
            def yaz():
                puanim = f'{entr.get()}'
                for u in range(1,aktif.max_row+1):
                    isimler = aktif[f"A{u}"].value.upper()
                    soy_isim = aktif[f"B{u}"].value.upper()
                    ogr_sinif = aktif[f"D{u}"].value.upper()
                    if isimler+" "+soy_isim == secilen_ogr and ogr_sinif == a.sinif:
                        a.puan = puanim[1:len(puanim)]
                        aktif.cell(column=3,row=u,value=puanim)
                        oku.save("ogr.xlsx")
                        wd.destroy()
            button = tk.Button(wd,text="KAYDET",font="bold 12",command=yaz).place(x=70,y=100)
            break
                        

ogr_bilgi_gor = tk.Button(window,text="bilgi".upper(),font="bold 14",width=12,command=bilgi).place(x=210,y=200)

ogrenci_list = tk.Listbox(window,bg="white",width=19,height=9,font="calibri 12")
kaydir = ttk.Scrollbar(window,command=ogrenci_list.yview)
kaydir.pack(side="right", fill="y")
ogrenci_list.config(yscrollcommand=kaydir.set)
ogrenci_list.place(x=200,y=10)
duzen = tk.Button(window,text="DUZENLE",font=("Aharoni",12,"bold"),command=duzenler).place(x=10,y=180)
window.mainloop()
